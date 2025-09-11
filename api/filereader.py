import csv, re
from openpyxl import load_workbook


EXTENTION_LIST = [
    "text/csv",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
]

FOUND = "Segment_found"


def read_setup_file(setup_file):
    """
    - Reads different types of files : txt, csv, xls
    - Loads rules from the file
    - returns a list of dictionaries
    """
    if setup_file.content_type not in EXTENTION_LIST:
        return None
    
    if setup_file.content_type == "text/csv":

        return read_csv_file(setup_file)
    
    if setup_file.content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
        return read_excel_file(setup_file)
    

def read_csv_file(setup_file):
    """
    - Reads csv file and segrigates the text and appends to empty list in a dictionary format.
    - Remives the first row as it assumes it as a header.
    """
    pre_defined_rules = []
    decoded_file = setup_file.read().decode("utf-8").splitlines()
    file_content = csv.reader(decoded_file)

    next(file_content, None)

    for col in file_content:
        if len(col) >=3:
            pre_defined_rules.append({
                "start": col[0].strip(),
                "end": col[1].strip(),
                "phrase": col[2].strip(),
                "segment_type": col[3].strip().lower() if len(col) >= 4 and col[3].strip() else "html"
            })

    return pre_defined_rules

def read_excel_file(setup_file):
    """
    - Reads excel file and segrigates the text and appends to empty list in a dictionary format.
    """
    pre_defined_rules = []
    work_book = load_workbook(setup_file, data_only=True)
    file_content = work_book.active

    for col in file_content.iter_rows(min_row = 2,values_only=True):
        if len(col) >= 3:    
            pre_defined_rules.append({
                    "start": col[0].strip(),
                    "end": col[1].strip() if col[1] else "",
                    "phrase": col[2].strip(),
                    "segment_type": str(col[3]).strip().lower() if len(col) >= 4 and col[3] else "html" 
                })

    return pre_defined_rules
        

def read_email_file(email_file):
    """
    - reads email file (.eml) into a list of lines with the line numbers.
    # """
    if not email_file.content_type == "message/rfc822":
        return None
    file_content=email_file.read().decode("utf-8", errors="ignore")
    lines = file_content.splitlines()
    return list(enumerate(lines, start=1))


def count_phrase(phrase, text):
    """
    - returns all the matched phrase and lines.
    """
    return re.findall(re.escape(phrase), text, flags=re.IGNORECASE)

def find_occurrences(lines, phrase):
    """
    - returns the lines where phrase was found and also the total count
    """
    occurrences = []
    total = 0
    for line_no, text in lines:
        matches = count_phrase(phrase, text)
        if matches:
            occurrences.append({
                "line": line_no,
                "text": text
            })
            total += len(matches)
    return occurrences, total

def segment_extraction(email_lines, rules):
    """
    - Scans the provided email_lines and then extracts the sgements based on all rules.
    - Returns the result in a dictionary format {}
      structure:
      "phrase" : {  
        "segment_lines":[]
        "start_line": integer,
        "occurrence": [{
        }],
        "total_count": integer
      }
    """

    results = {r["phrase"]: {
        "segment_lines": [],
        "start_line": None,
        "status": "segment_not_found",
        "matched_segments": [],
        "total_count": 0
    } for r in rules}

    active = {}

    for line_no, text in email_lines:
        for rule in rules:
            start, end = rule.get("start", ""), rule.get("end", "")
            phrase = rule["phrase"]
            seg_type = rule.get("segment_type", "single").lower()

            # --- Single line segment ---
            if seg_type == "single" and start in text:
                occ, total = find_occurrences([(line_no, text)], phrase)
                results[phrase] = {
                    "segment_lines": [text],
                    "start_line": line_no,
                    "status": FOUND,
                    "matched_segments": occ,
                    "total_count": total
                }

            # --- Multi-line / HTML segment ---
            elif seg_type != "single":
                if start in text and phrase not in active:
                    active[phrase] = {"rule": rule, "lines": [(line_no, text)], "start_line": line_no}
                elif phrase in active:
                    active[phrase]["lines"].append((line_no, text))
                    if end and end in text:
                        lines = active[phrase]["lines"]
                        occ, total = find_occurrences(lines, phrase)
                        results[phrase] = {
                            "segment_lines": [t for _, t in lines],
                            "start_line": active[phrase]["start_line"],
                            "status": FOUND,
                            "matched_segments": occ,
                            "total_count": total
                        }
                        del active[phrase]

    # --- Any still-open multi-line segments ---
    for phrase, state in active.items():
        lines = state["lines"]
        occ, total = find_occurrences(lines, phrase)
        results[phrase] = {
            "segment_lines": [t for _, t in lines],
            "start_line": state["start_line"],
            "status": FOUND,
            "matched_segments": occ,
            "total_count": total
        }
    
    if check_total_count(results) is None:
        return None
    return results

def check_total_count(result):
    """
    - checks if each total_count in the provided data is 0
    """
    if all(item.get("total_count", 0) == 0 for item in result.values()):
        return None
    
    return result

